"""
storage/excel_logger.py
=======================
Zapisywanie pomiarów do pliku Excel (.xlsx) z wyrównaniem do prawej.

Każda sesja rejestracji dopisuje dane do istniejącego pliku (jeśli istnieje)
albo tworzy nowy z nagłówkiem. Po każdym wierszu zapisujemy plik na dysk —
chroni to przed utratą danych przy crashu aplikacji.
"""

import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import TIMESTAMP_FORMAT

# Szerokości kolumn w jednostkach Excela (1 jednostka ≈ szerokość znaku '0').
_COLUMN_WIDTHS = {
    1: 22,   # Czas
    2: 14,   # Masa
    3: 10,   # Jednostka
    4: 12,   # Status
}

_RIGHT = Alignment(horizontal="right", vertical="center")
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")  # jasnoniebieski


class ExcelLogger:
    """Logger Excel. Otwiera plik .xlsx, dopisuje wiersze, zapisuje po każdym pomiarze."""

    HEADERS = ["Czas", "Masa", "Jednostka", "Status"]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._wb = None
        self._ws = None

    def open(self) -> None:
        """Otwiera lub tworzy plik .xlsx. Jeśli plik istnieje — dopisujemy do niego."""
        if os.path.exists(self.file_path) and os.path.getsize(self.file_path) > 0:
            self._wb = openpyxl.load_workbook(self.file_path)
            self._ws = self._wb.active
        else:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._ws.title = "Pomiary"
            self._write_header()
            self._wb.save(self.file_path)

    def write_row(
        self,
        timestamp: datetime,
        mass: Optional[float],
        unit: str,
        status: str,
    ) -> None:
        """Dopisuje jeden wiersz pomiaru i zapisuje plik."""
        if self._ws is None:
            raise RuntimeError("ExcelLogger nie został otwarty — wywołaj open() najpierw.")

        mass_val = None if mass is None else round(mass, 6)
        row_data = [timestamp.strftime(TIMESTAMP_FORMAT), mass_val, unit, status]

        self._ws.append(row_data)

        # Wyrównanie do prawej dla każdej komórki nowego wiersza.
        row_idx = self._ws.max_row
        for col in range(1, len(row_data) + 1):
            self._ws.cell(row=row_idx, column=col).alignment = _RIGHT

        self._wb.save(self.file_path)

    def close(self) -> None:
        """Zamyka logger. Bezpieczne wielokrotne wywołanie."""
        self._wb = None
        self._ws = None

    def _write_header(self) -> None:
        """Wpisuje nagłówek z formatowaniem i ustawia szerokości kolumn."""
        self._ws.append(self.HEADERS)

        for col, header in enumerate(self.HEADERS, start=1):
            cell = self._ws.cell(row=1, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _RIGHT
            self._ws.column_dimensions[get_column_letter(col)].width = _COLUMN_WIDTHS.get(col, 14)

    # Context manager
    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False
